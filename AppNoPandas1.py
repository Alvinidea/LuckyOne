from datetime import datetime
import tkinter as tk
import time
import tkinter.messagebox
import random
import threading
# import pandas as pd
import json
import os
import openpyxl
from openpyxl import load_workbook
'''
需要初始文件
    groups.xlsx
    finished.xlsx
    missions.xlsx
'''
LARGE_FONT = ("Verdana", 12)
'''
GROUPS = "config/groups.xlsx"
MISSIONS = "config/missions.xlsx"
FINISHED = "config/finished.xlsx"
MATCH = "config/match.xlsx"

配置文件： config.txt
{
"groups":"groups.xlsx",
"match":"match.xlsx",
"mission":"missions.xlsx",
"finished":"finished.xlsx",
"classlist":"classlist.xlsx"
}
'''
# 加载配置信息
def init():
    global config
    global GROUPS, MISSIONS, FINISHED, MATCH, WIDTH, ClASSLIST, HEIGHT
    with open("config/config.txt", 'r', encoding='utf-8') as load_f:
        config = "config/"
        load_dict = json.load(load_f)
        GROUPS = config + load_dict['groups']
        MISSIONS = config + load_dict['mission']
        FINISHED = config + load_dict['finished']
        MATCH = config + load_dict['match']
        # 班级名单
        ClASSLIST = config + load_dict['classlist']

    WIDTH = 500
    HEIGHT = 450
    # 判断 3个文件是否存在
    if os.path.exists(GROUPS) == False:
        tkinter.messagebox.showwarning(title='Warn', message="分组信息不存在！请在config文件夹中导入分组信息！")
    if os.path.exists(MISSIONS) == False:
        tkinter.messagebox.showwarning(title='Warn', message="题目信息不存在！请在config文件夹中导入题目信息！")
    if os.path.exists(ClASSLIST) == False:
        tkinter.messagebox.showwarning(title='Warn', message="班级信息不存在！请在config文件夹中导入班级信息！")
    if os.path.exists(FINISHED) == False:
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.append(['组号', '题号', '成员', '时间'])
        wb.save(FINISHED)

# excel的操作类
class excel():
    def __init__(self, file, create=True):
        self.file = file
        if create == True:
            self.wb = load_workbook(self.file)
            sheets = self.wb.sheetnames
            self.sheet = sheets[0]
            self.ws = self.wb[self.sheet]
        else:
            self.wb = openpyxl.Workbook()
            sheets = self.wb.sheetnames
            self.sheet = sheets[0]
            self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns


    def getRowsNum(self):
        rows = self.ws.max_row
        return rows


    def getColsNum(self):
        columns = self.ws.max_column
        return columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    # 获取某行所有值
    def addRow(self, row):
        try:
            self.ws.append(row)
            self.wb.save(self.file)
        except:
            self.ws.append([])
            self.wb.save(self.file)

    # 跟据某列的值获取行数据
    def getRowByColVulue(self, column, value):
        cols = self.getColValues(column)
        for irow in range(1, self.getRowsNum()+1):
            if value == self.getCellValue(irow, column):
                print(self.getRowValues(irow))
                return self.getRowValues(irow)


class Application(tk.Tk):
    '''
    多页面测试程序
        界面与逻辑分离
    '''
    def __init__(self):
        super().__init__()

        self.title("Fate")
        self.geometry('500x450')

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        menubar = tk.Menu(self)
        filemenu = tk.Menu(menubar, tearoff=0)
        newmenu = tk.Menu(menubar, tearoff=0)

        menubar.add_cascade(label='决定命运', menu=filemenu)
        menubar.add_cascade(label='新学期', menu=newmenu)

        filemenu.add_command(label="抽签选组", command=self.chouqian)
        filemenu.add_command(label="分配题目", command=self.xuanti)
        filemenu.add_command(label="随机抽人", command=self.luckypeo)

        newmenu.add_command(label="选择班级", command=self.selectClass)
        newmenu.add_command(label="新学期", command=self.xinxueqi)
        newmenu.add_command(label="刷新", command=self.flush)
        self.config(menu=menubar)
        self.frames = {}
        for F in (StartPage_Sign, Pagetwo_Match, PageThree_New, PageFour_Luc, PageFive_selectClass):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
            # 四个页面的位置都是 grid(row=0, column=0), 位置重叠，只有最上面的可见！！
        self.show_frame(StartPage_Sign)


    # 展示框架
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()  # 切换，提升当前 tk.Frame z轴顺序（使可见）！！此语句是本程序的点睛之处


    def chouqian(self,):
        self.show_frame(StartPage_Sign)


    def xuanti(self):
        self.show_frame(Pagetwo_Match)


    def xinxueqi(self):
        self.show_frame(PageThree_New)

    def selectClass(self):
        self.show_frame(PageFive_selectClass)


    def luckypeo(self):
        self.show_frame(PageFour_Luc)


    def flush(self):
        for F in (StartPage_Sign, Pagetwo_Match, PageThree_New):
            frame = self.frames[F]
            frame.flush()


class StartPage_Sign(tk.Frame):
    '''主1页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.label = tk.Label(self, text="谁是幸运儿！！！", font=LARGE_FONT,bg='green', fg='white')
        # self.label.pack(pady=10, padx=10)
        self.label.place(x=WIDTH/2-100, y=10, width=200, height=25)
        # -------------------------------------------- 选组
        self.LabelsFrame1 = tk.LabelFrame(self, text="哪组会中奖？", padx=35, pady=35)
        # self.LabelsFrame1.pack()
        self.LabelsFrame1.place(x=20, y=50, width=WIDTH/2 -30, height=HEIGHT-200)
        self.first = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 20, "normal"))
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 30, "normal"))
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 20, "normal"))
        self.third.pack()

        self.frame1 = tk.Frame(self.LabelsFrame1, width=120, height=50, pady=10, padx=10)
        self.frame1.pack()
        self.btnStart = tk.Button(self.frame1, text='开始', bg='green', fg='white', command=self.butStartClick)
        self.btnStart.place(x=0, y=5, width=45, height=25)
        self.butStop = tk.Button(self.frame1, text='停止',  bg='green', fg='white', command=self.btnStopClick)
        self.butStop.place(x=50, y=5, width=45, height=25)
        # -------------------------------------------- 选人
        self.LabelsFrame2 = tk.LabelFrame(self, text="哪位同学展示？", padx=35, pady=35)
        # self.LabelsFrame2.pack()
        self.LabelsFrame2.place(x=WIDTH/2+20, y=50, width=WIDTH/2 -30, height=HEIGHT-200)
        self.first2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 20, "normal"))
        self.first2.pack()

        self.second2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 30, "normal"))
        self.second2['fg'] = 'red'
        self.second2.pack()

        self.third2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 20, "normal"))
        self.third2.pack()
        self.frame2 = tk.Frame(self.LabelsFrame2, width=120, height=50, pady=10, padx=10)
        self.frame2.pack()
        self.btnStart2 = tk.Button(self.frame2, text='开始', bg='green', fg='white', command=self.butStartClick2)
        self.btnStart2.place(x=0, y=5, width=45, height=25)
        self.butStop2 = tk.Button(self.frame2,  text='停止',  bg='green', fg='white', command=self.btnStopClick2)
        self.butStop2.place(x=50, y=5, width=45, height=25)
        # LabelFrame
        self.speed = 0.02           # 滚动速度
        self.Groupflag = True
        self.Memberflag = True
        self.LuckyGroup = 0
        self.LuckyMember = ''
        self.GroupInfo, self.Groups, self.Finished = self.importInfo()
        self.differenceGroup = list(set(self.Groups).difference(set(self.Finished)))
        self.Members = []


    # 初始导入组 和 已选组的信息
    def importInfo(self):
        # 导入所有组
        exc = excel(GROUPS)
        grow, gcol = exc.getRowsClosNum()
        groups = {}
        for index in range(1, grow+1):
            tempG = exc.getRowValues(index)
            groups[tempG[0]] = tempG[0:]
        group = exc.getColValues(1)[1:]


        fexc = excel(FINISHED)
        frow, fcol = fexc.getRowsClosNum()
        retFinished = list()
        if grow <= 1:
            tkinter.messagebox.showwarning(title='提示', message='\"' + GROUPS+' \"分组文件中没有数据！！！')
        if frow > 1:
            retFinished = fexc.getColValues(1)[1:]
        return groups, group, retFinished


    # 保存抽取的组 和 组员信息
    def saveSelect(self):
        wb = openpyxl.load_workbook(FINISHED)
        # 获取某一特定的工作表
        sheet = wb["Sheet"]

        exc = excel(MATCH)
        # 2 : 根据 组号（第一列）的值  ，获取该列所在的行
        question = exc.getRowByColVulue(1, self.LuckyGroup)[1]
        # question = match[match['组号'] == self.LuckyGroup].values.tolist()[0][2]
        now = datetime.now()
        strnow = datetime.strftime(now, '%Y-%m-%d %H:%M:%S')
        print([self.LuckyGroup, question, self.LuckyMember, strnow])

        sheet.append([self.LuckyGroup, question, self.LuckyMember, strnow])
        wb.save(FINISHED)


    # 抽取完成之后进行更新
    def updateDiff(self):
        self.GroupInfo, self.Groups, self.Finished = self.importInfo()
        self.differenceGroup = list(set(self.Groups).difference(set(self.Finished)))


    # 选择 组 的滚动条
    def switch(self):
        self.Groupflag = True
        while self.Groupflag:
            i = random.randint(0, len(self.Groups)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            self.third['text'] = 'Group' + str(self.Groups[i])
            time.sleep(self.speed)


    # 选择 成员 的滚动条
    def switch2(self):
        self.Memberflag = True
        while self.Memberflag:
            i = random.randint(0, len(self.Members)-1)
            self.first2['text'] = self.second2['text']
            self.second2['text'] = self.third2['text']
            self.third2['text'] = str(self.Members[i])
            time.sleep(self.speed)
        # 选组号 开始按钮


    def butStartClick(self):
        t = threading.Thread(target=self.switch)
        t.start()

        # 选组号 结束按钮


    def btnStopClick(self):
        if self.Groupflag == True:
            self.Groupflag = False
            # 随机获取组号
            if len(self.differenceGroup) == 0:
                tkinter.messagebox.showwarning("提示", "所有的组都已经完成！！！")
                return
            index = random.randint(0, len(self.differenceGroup) - 1)
            self.second['text'] = "Group" + str(self.differenceGroup[index])
            self.LuckyGroup = self.differenceGroup[index]
            # 获取组号对应的成员
            # members = self.GroupInfo[self.GroupInfo['组号'] == self.LuckyGroup].values.tolist()[0][2:]
            print(self.GroupInfo, self.LuckyGroup)
            members = self.GroupInfo[self.LuckyGroup][3:]
            self.Members = members


    # 选成员 开始按钮
    def butStartClick2(self):
        if len(self.Members) != 0:
            t = threading.Thread(target=self.switch2)
            t.start()
        else:
            tkinter.messagebox.showwarning("提示", "请先抽出选择的组！！！")


    # 选成员 结束按钮
    def btnStopClick2(self):
        if self.Memberflag == True:
            self.Memberflag = False
            # self.LuckyMember = self.second2['text']
            index = random.randint(0, len(self.Members)-1)
            self.second2['text'] = self.Members[index]
            self.LuckyMember = self.Members[index]
            # print(self.LuckyMember)
            self.saveSelect()
            # 选一个组号和成员之后需要更新以为选择的组的记录
            self.updateDiff()

    # 刷新所有 flag
    def flush(self):
        self.Groupflag = False
        self.Memberflag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        self.Groupflag = True
        self.Memberflag = True


class Pagetwo_Match(tk.Frame):
    '''第2页'''

    def __init__(self, parent, root):
        super().__init__(parent)
        self.LabelsFrame = tk.LabelFrame(self, text="你们做什么？", padx=25, pady=20)
        self.LabelsFrame.place(x=20, y=35, width=WIDTH/2 -30, height=HEIGHT-100)

        self.first = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame , text='？？？', font=("宋体", 30, "normal"))
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame , text='？？？', font=("宋体", 20, "normal"))
        self.third.pack()


        self.first2 = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.first2.pack()
        self.second2 = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 30, "normal"))
        self.second2['fg'] = 'red'
        self.second2.pack()
        self.third2 = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.third2.pack()

        self.frameBtn = tk.Frame(self.LabelsFrame, width=200, height=50, pady=10, padx=10)
        self.frameBtn.pack()
        # self.frameBtn.place(x=WIDTH/2+20, y=50, width=WIDTH/2 -30, height=HEIGHT-100)

        self.btnStart = tk.Button(self.frameBtn, text='开始', fg='white', bg='green', command=self.butStartClick)
        self.btnStart.place(x=0, y=5, width=45, height=25)
        self.btnEnd = tk.Button(self.frameBtn, text='结束', fg='white', bg='green', command=self.butEndClick)
        self.btnEnd.place(x=50, y=5, width=45, height=25)
        self.btnSave = tk.Button(self.frameBtn, text='SAVE', fg='white', bg='green', command=self.SaveClick)
        self.btnSave.place(x=100, y=5, width=45, height=25)

        self.lframeText = tk.LabelFrame(self, text='匹配信息', pady=10, padx=10)
        self.lframeText.place(x=WIDTH / 2 + 20, y=35, width=WIDTH / 2 - 30, height=HEIGHT - 100)
        self.text = tk.Text(self.lframeText)
        self.text.place(x=0, y=2, width=WIDTH/2-55, height=HEIGHT-150)

        self.speed = 0.02  # 滚动速度
        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions = self.matchInfo()
        # 防止 开始按钮 多次点击
        self.startMatchFlag = True
        # 匹配前控制 滚动条 的滚动信号
        self.flag = True
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = True
        # 匹配时候禁止点击 停止 按钮
        self.showflag = True

        # 组号 --- 题号
        self.Match = {}


    def matchInfo(self):
        # 导入所有组
        exc = excel(GROUPS)
        grow, gcol = exc.getRowsClosNum()
        groups = {}
        for index in range(1, grow+1):
            tempG = exc.getRowValues(index)
            groups[tempG[0]] = tempG[0:]
        group = exc.getColValues(1)[1:]

        mexc = excel(MISSIONS)
        mrow, mcol = mexc.getRowsClosNum()
        missions = {}
        for index in range(1, mrow + 1):
            tempM = mexc.getRowValues(index)
            missions[tempM[0]] = tempM[0:]
        mission = exc.getColValues(1)[1:]

        if mrow == 0:
            tkinter.messagebox.showwarning("提示", '\"'+MISSIONS+"\" 题目文件中没有信息！")

        random.shuffle(group)
        random.shuffle(mission)

        return groups, missions, group, mission


    # 选择 组/题 的滚动条
    def switch(self):
        # self.flag = True
        while self.flag:
            i = random.randint(0, len(self.Missions)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            self.third['text'] = 'Mission' + str(self.Missions[i])

            j = random.randint(0, len(self.Groups) - 1)
            self.first2['text'] = self.second2['text']
            self.second2['text'] = self.third2['text']
            self.third2['text'] = 'Group' + str(self.Groups[j])
            time.sleep(self.speed)
            if( self.flag == False):
                break


    def show(self):
        if self.showflag == True:
            # 防止多次点击 结束 按钮
            self.showflag = False
            self.text.insert('end', '\n' + "组号\t|\t题号" )
            self.text.insert('end', '\n--------------------------')
            for gro in self.Match.keys():
                self.second['text'] = "组号" + str(gro)
                self.second2['text'] = "题号" + str(self.Match[gro])
                self.text.insert('end', '\n' + str(gro) + "\t|\t" + str(self.Match[gro]))
                time.sleep(0.8)
            self.text.insert('end', '\n\n\t 分组数目：' + str(len(self.Match.keys())))
            # 控制匹配的时候 的滚动
            self.fflag = False
            tkinter.messagebox.showinfo(title="提示", message="匹配完成QAQ")


    def switch2(self):
        # self.fflag = True
        while self.fflag:
            i = random.randint(0, len(self.Missions)-1)
            self.first['text'] = self.third['text']
            # self.second['text'] = self.third['text']
            self.third['text'] = 'Mission' + str(self.Missions[i])

            j = random.randint(0, len(self.Groups) - 1)
            self.first2['text'] = self.third2['text']
            # self.second2['text'] = self.third2['text']
            self.third2['text'] = 'Group' + str(self.Groups[j])
            time.sleep(self.speed)
            if( self.fflag == False):
                break


    # 选组号 开始按钮
    def butStartClick(self):
        if(self.startMatchFlag == True):
            t = threading.Thread(target=self.switch)
            t.start()
            self.startMatchFlag = False

        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions = self.matchInfo()
        for i in range(0, len(self.Groups)):
            self.Match[self.Groups[i]] = self.Missions[i % len(self.Missions)]


    def butEndClick(self):
        self.flag = False
        t = threading.Thread(target=self.show)
        t.start()
        t2 = threading.Thread(target=self.switch2)
        t2.start()


    # 保存匹配结果
    def SaveClick(self):
        if self.flag != False:
            tkinter.messagebox.showinfo(title="提示", message="您还没有进行匹配QAQ")
            return
        if tkinter.messagebox.askyesno(title="提示",message="确定保存分配结果吗？") == True:
            self.save_MatchInfo()


    def save_MatchInfo(self):
        exc = excel(MATCH, False)
        row = ['组号', '题号', '题目', '成员1', '成员2', '成员3', '成员4', '备注']
        exc.addRow(row)
        # print(self.MissionsInfo)
        for Gnum in range(0, len(self.Groups)):
            name = self.MissionsInfo[self.Missions[Gnum]][1]
            members = self.GroupsInfo[self.Groups[Gnum]][2:]
            newRow = [self.Groups[Gnum], self.Missions[Gnum], name,
                      members[0], members[1], members[2], members[3],
                      ""]
            exc.addRow(newRow)
        tkinter.messagebox.showinfo(title="提示", message="保存成功")


    def flush(self):
        # 防止 开始按钮 多次点击
        self.startMatchFlag = False
        # 匹配前控制 滚动条 的滚动信号
        self.flag = False
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = False
        # 匹配时候禁止点击 停止 按钮
        self.showflag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        # ===================================
        # 防止 开始按钮 多次点击
        self.startMatchFlag = True
        # 匹配前控制 滚动条 的滚动信号
        self.flag = True
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = True
        # 匹配时候禁止点击 停止 按钮
        self.showflag = True

        self.text.delete(0.0, tkinter.END)


class PageThree_New(tk.Frame):
    '''第3页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        """        
        self.label = tk.Label(self, text="这是第二页", font=LARGE_FONT)
        self.label.pack(pady=10, padx=10)
        self.grlabel = tk.Label(self, text="分组文件名", font=LARGE_FONT)
        self.grlabel.pack()
        self.grentry = tk.Entry(self, show=None, font=('Arial', 14))
        self.grentry.pack()

        self.missionlabel = tk.Label(self, text="题目文件名", font=LARGE_FONT)
        self.missionlabel.pack()
        self.missionentry = tk.Entry(self, show=None, font=('Arial', 14))
        self.missionentry.pack()
        

        self.matchlabel = tk.Label(self, text="匹配文件名", font=LARGE_FONT)
        self.matchlabel.pack()
        self.matchentry = tk.Entry(self, show=None, font=('Arial', 14))
        self.matchentry.pack()
        """
        list_label = self.init()
        self.LabelFrame = tk.LabelFrame(self, text="配置文件信息", pady=50, width=WIDTH-100, height=HEIGHT-180)
        self.LabelFrame.pack()
        self.grlabel = tk.Label(self.LabelFrame, text=list_label[0], font=("宋体", 16, "normal"))
        self.grlabel.pack()
        self.missionlabel = tk.Label(self.LabelFrame, text=list_label[1], font=("宋体", 16, "normal"))
        self.missionlabel.pack()
        self.matchlabel = tk.Label(self.LabelFrame, text=list_label[2], font=("宋体", 16, "normal"))
        self.matchlabel.pack()
        self.finishedlabel = tk.Label(self.LabelFrame, text=list_label[3], font=("宋体", 16, "normal"))
        self.finishedlabel.pack()
        self.entryClassName = tk.Entry(self)
        self.entryClassName.pack(padx=50, pady=5)
        self.Confrimbutton = tk.Button(self, text="点击确认开始新学期", fg='white', bg='green', command=self.confrim)
        self.Confrimbutton.pack(padx=50, pady=50)



    def init(self):
        with open("config/config.txt", 'r', encoding='utf-8') as load_f:
            config = "config/"
            listc = []
            load_dict = json.load(load_f)
            print(load_dict)
            listc.append(config + load_dict['groups'])
            listc.append(config + load_dict['mission'])
            listc.append(config + load_dict['finished'])
            listc.append(config + load_dict['match'])
            # 班级名单
            ClASSLIST = config + load_dict['classlist']
            return listc


    def confrim(self):
        if tkinter.messagebox.askyesno(title='提示', message="确定创建文件！开始新学期吗？") == True:
            className = self.entryClassName.get()
            if className == "":
                tkinter.messagebox.showwarning(title="提示", message="请输入班级名称！")
            else:
                self.modify()
                tkinter.messagebox.showinfo(title='提示', message='新学期所需的文件groups、missions、finished已经创建在config文件夹中！')
        else:
            tkinter.messagebox.showerror(title='Hi', message='取消操作')


    def getTimeInfo(self):
        nowt = datetime.now()
        timeInfo = datetime.strftime(nowt, '%Y-%m-%d')
        return timeInfo

    # 创建新班级的配置文件
    # config/ className / config file
    def modify(self):
        config = "config/"
        # className = "Test"
        className = self.entryClassName.get()
        if os.path.exists(config+className) == True:
            tkinter.messagebox.showwarning(title="提示", message=className+" 文件夹已经存在，请重新输入班级名称！")
            return
        # 创建文件夹
        os.makedirs(config+className)
        timeInfo = self.getTimeInfo()
        groups = className+'/groups'+timeInfo+".xlsx"
        missions = className+'/missions'+timeInfo+".xlsx"
        finished = className+'/finished'+timeInfo+".xlsx"
        classlist = className+'/classlist' + timeInfo + ".xlsx"
        configinfo = {}
        configinfo['groups'] = groups
        configinfo['mission'] = missions
        configinfo['finished'] = finished
        configinfo['classlist'] = classlist
        configinfo['match'] = className+'/match.xlsx'
        with open("config/config.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict['groups'] = groups
            load_dict['mission'] = missions
            load_dict['finished'] = finished
            load_dict['classlist'] = classlist
            str = json.dumps(load_dict)
            load_f.seek(0)
            load_f.truncate()
            load_f.write(str)
        # ==================================
        wb1 = openpyxl.Workbook()
        sheet = wb1['Sheet']
        sheet.append(['组号', '组名', '成员1', '成员2', '成员3', '成员4'])
        wb1.save(config+groups)
        # ==================================
        wb2 = openpyxl.Workbook()
        sheet2 = wb2['Sheet']
        sheet2.append(['组号', '题号', '成员', '时间'])
        wb2.save(config+finished)
        # ==================================
        wb3 = openpyxl.Workbook()
        sheet3 = wb3['Sheet']
        sheet3.append(['题号', '题目', '要求', '备注'])
        wb3.save(config+missions)

        self.addConfigToClassConfig(className, configinfo)


    def addConfigToClassConfig(self, configName,configInfo):
        with open("config/ClassConfig.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict[configName] = configInfo
            str = json.dumps(load_dict)
            load_f.seek(0)
            # 清空之前的信息
            load_f.truncate()
            load_f.write(str)



    def flush(self):
        pass


class PageFour_Luc(tk.Frame):
    '''第4页'''
    def __init__(self, parent, root):
        super().__init__(parent)

        self.LabelsFrame = tk.LabelFrame(self, text="谁是幸运星", padx=25, pady=20)
        self.LabelsFrame.place(x=50, y=50, width=WIDTH-120, height=HEIGHT-100)
        self.first = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 40, "normal"))
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.third.pack()
        self.Confrimbutton = tk.Button(self.LabelsFrame, text="幸运星", fg='white', bg='green', command=self.ramdom_peo)
        self.Confrimbutton.pack(padx=25, pady=10)
        self.Repeatbutton = tk.Button(self.LabelsFrame, text="重选", fg='white', bg='green', command=self.repeat)
        self.Repeatbutton.pack(padx=25, pady=10)

        self.speed = 0.02
        self.flag = True
        self.ClassInfo = self.initClassInfo()
        self.NameList = self.getNameList()


    def initClassInfo(self):
        if os.path.exists(ClASSLIST) == False:
            tkinter.messagebox.showwarning(title='提示！',message=ClASSLIST+"文件不存在，无法实现单人点名操作" )
            return dict()
        exc = excel(ClASSLIST)
        rown = exc.getColsNum()
        classInfo = {}
        for col in range(1, rown+1):
            tempCol = exc.getColValues(col)
            classInfo[tempCol[0]] = tempCol[1:]
        return classInfo


    def getNameList(self):
        for colName in self.ClassInfo.keys():
            if colName == "Name" or colName == "姓名" or colName=="name":
                return self.ClassInfo[colName]
        else:
            return list()


    # 选择 组 的滚动条
    def switch(self):
        while self.flag:
            i = random.randint(0, len(self.NameList)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            self.third['text'] =  str(self.NameList[i])
            time.sleep(self.speed)
            if self == False:
                break

    def threadingStop(self):
        time.sleep(3)
        self.flag = False

    def ramdom_peo(self):
        t = threading.Thread(target=self.switch)
        t.start()
        t1 = threading.Thread(target=self.threadingStop)
        t1.start()


    def repeat(self):
        self.flag = True
        self.ramdom_peo()


    def flush(self):
        self.flag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        self.flag = True
        pass


"""
班级信息配置分为默认配置 和 选择配置

默认配置的信息：config.txt
选择配置的信息：ClassConfig.txt
"""


class PageFive_selectClass(tk.Frame):
    '''第4页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.LabelsFrame = tk.LabelFrame(self, text="班级选择", padx=25, pady=20)
        self.LabelsFrame.place(x=50, y=50, width=WIDTH-120, height=HEIGHT-100)
        self.classes, self.configs = self.initClassInfo()
        self.Lebels = list()
        self.Buttons = list()
        self.frames = list()
        for clas in self.classes:
            frame = tk.Frame(self.LabelsFrame, width=WIDTH-50, height=50)
            label = tk.Label(frame, text=clas, font=LARGE_FONT)
            button = tk.Button(frame, text="选择"+clas, font=LARGE_FONT,
                               command=lambda arg=clas:self.selectClass(arg))

            frame.pack(padx=20, pady=8)
            label.place(x=5, y=10)
            button.place(x=100, y=10)
            self.frames.append(frame)
            self.Lebels.append(label)
            self.Buttons.append(button)


    def initClassInfo(self):
        with open("config/ClassConfig.txt", 'r', encoding='utf-8') as load_f:
            config = "config/"
            # load_dict = json.load(load_f, strict = False)
            load_dict = json.load(load_f)
            # classes = load_dict.keys()
            # configs = load_dict
            return load_dict.keys(), load_dict


    def getConfigByClass(self, cls):
        return self.configs[cls]


    def reSetMemory(self, config):
        configPath = "config/"
        GROUPS = configPath+config['groups']
        MISSIONS = configPath+config['mission']
        FINISHED = configPath+config['finished']
        MATCH = configPath+config['match']
        # 班级名单
        ClASSLIST = configPath+config['classlist']


    def reSetConfig(self, config):
        self.reSetMemory(config)
        with open("config/config.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict['groups'] = config['groups']
            load_dict['mission'] = config['mission']
            load_dict['finished'] = config['finished']
            load_dict['match'] = config['match']
            load_dict['classlist'] = config['classlist']
            str = json.dumps(load_dict)
            load_f.seek(0)
            load_f.truncate()
            load_f.write(str)


    def selectClass(self, cls):
        # 根据班名获取班级配置信息
        config = self.getConfigByClass(cls)
        # 重新设置默认配置信息
        self.reSetConfig(config)
        tkinter.messagebox.showinfo(title='操作成功！',
                                    message="班级配置信息已经重新设置！\n当前班级为："+cls)



if __name__ == '__main__':
    # 初始化配置
    init()
    # 实例化Application
    app = Application()
    # 主消息循环:
    app.mainloop()