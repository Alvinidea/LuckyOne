from datetime import datetime
import tkinter as tk
import tkinter.messagebox
import random
import threading
import pandas as pd
from tkinter import ttk
import time
import json
import os
import openpyxl
'''
需要初始文件
    groups.xlsx
    finished.xlsx
    missions.xlsx
'''
LARGE_FONT = ("Verdana", 12)
'''
GROUPS = "config/groups.xlsx"
MISSIONS = "config/missions.xlsx"
FINISHED = "config/finished.xlsx"
MATCH = "config/match.xlsx"

配置文件： config.txt
{
"groups":"groups.xlsx",
"match":"match.xlsx",
"mission":"missions.xlsx",
"finished":"finished.xlsx"
}
'''
# 加载配置信息
def init():
    global GROUPS, MISSIONS, FINISHED, MATCH, WIDTH, HEIGHT
    with open("config/config.txt", 'r', encoding='utf-8') as load_f:
        # load_dict = json.load(load_f)[0]
        load_dict = json.load(load_f)
        GROUPS = "config/" + load_dict['groups']
        MISSIONS = "config/" + load_dict['mission']
        FINISHED = "config/" + load_dict['finished']
        MATCH = "config/" + load_dict['match']

        WIDTH = 500
        HEIGHT = 450
        # 判断 3个文件是否存在
        if os.path.exists(GROUPS) == False:
            tkinter.messagebox.showwarning(title='Warn', message="分组信息不存在！请在config文件夹中导入分组信息！")
        if os.path.exists(MISSIONS) == False:
            tkinter.messagebox.showwarning(title='Warn', message="题目信息不存在！请在config文件夹中导入题目信息！")
        if os.path.exists(FINISHED) == False:
            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet.append(['组号', '题号', '成员', '时间'])
            wb.save(FINISHED)


class Application(tk.Tk):
    '''
    多页面测试程序
        界面与逻辑分离
    '''
    def __init__(self):
        super().__init__()

        self.title("Fate")
        self.geometry('500x450')

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        menubar = tk.Menu(self)
        filemenu = tk.Menu(menubar, tearoff=0)
        newmenu = tk.Menu(menubar, tearoff=0)

        menubar.add_cascade(label='决定命运', menu=filemenu)
        menubar.add_cascade(label='新学期', menu=newmenu)

        filemenu.add_command(label="抽签", command=self.chouqian)
        filemenu.add_command(label="选题", command=self.xuanti)
        filemenu.add_command(label="幸运星", command=self.luckypeo)
        newmenu.add_command(label="新学期", command=self.xinxueqi)
        newmenu.add_command(label="刷新", command=self.flush)
        self.config(menu=menubar)
        self.frames = {}
        for F in (StartPage_Sign, Pagetwo_Match, PageThree_New, PageFour_Luc):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
            # 四个页面的位置都是 grid(row=0, column=0), 位置重叠，只有最上面的可见！！
        self.show_frame(StartPage_Sign)


    # 展示框架
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()  # 切换，提升当前 tk.Frame z轴顺序（使可见）！！此语句是本程序的点睛之处


    def chouqian(self,):
        self.show_frame(StartPage_Sign)


    def xuanti(self):
        self.show_frame(Pagetwo_Match)


    def xinxueqi(self):
        self.show_frame(PageThree_New)


    def luckypeo(self):
        self.show_frame(PageFour_Luc)


    def flush(self):
        for F in (StartPage_Sign, Pagetwo_Match, PageThree_New):
            frame = self.frames[F]
            frame.flush()





class StartPage_Sign(tk.Frame):
    '''主1页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.label = tk.Label(self, text="谁是幸运儿！！！", font=LARGE_FONT,bg='green', fg='white')
        # self.label.pack(pady=10, padx=10)
        self.label.place(x=WIDTH/2-100, y=10, width=200, height=25)
        # -------------------------------------------- 选组
        self.LabelsFrame1 = tk.LabelFrame(self, text="哪组会中奖？", padx=35, pady=35)
        # self.LabelsFrame1.pack()
        self.LabelsFrame1.place(x=20, y=50, width=WIDTH/2 -30, height=HEIGHT-200)
        self.first = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 20, "normal"))
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 30, "normal"))
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 20, "normal"))
        self.third.pack()

        self.frame1 = tk.Frame(self.LabelsFrame1, width=120, height=50, pady=10, padx=10)
        self.frame1.pack()
        self.btnStart = tk.Button(self.frame1, text='开始', bg='green', fg='white', command=self.butStartClick)
        self.btnStart.place(x=0, y=5, width=45, height=25)
        self.butStop = tk.Button(self.frame1, text='停止',  bg='green', fg='white', command=self.btnStopClick)
        self.butStop.place(x=50, y=5, width=45, height=25)
        # -------------------------------------------- 选人
        self.LabelsFrame2 = tk.LabelFrame(self, text="哪位同学展示？", padx=35, pady=35)
        # self.LabelsFrame2.pack()
        self.LabelsFrame2.place(x=WIDTH/2+20, y=50, width=WIDTH/2 -30, height=HEIGHT-200)
        self.first2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 20, "normal"))
        self.first2.pack()

        self.second2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 30, "normal"))
        self.second2['fg'] = 'red'
        self.second2.pack()

        self.third2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 20, "normal"))
        self.third2.pack()
        self.frame2 = tk.Frame(self.LabelsFrame2, width=120, height=50, pady=10, padx=10)
        self.frame2.pack()
        self.btnStart2 = tk.Button(self.frame2, text='开始', bg='green', fg='white', command=self.butStartClick2)
        self.btnStart2.place(x=0, y=5, width=45, height=25)
        self.butStop2 = tk.Button(self.frame2,  text='停止',  bg='green', fg='white', command=self.btnStopClick2)
        self.butStop2.place(x=50, y=5, width=45, height=25)
        # LabelFrame
        self.speed = 0.02           # 滚动速度
        self.Groupflag = True
        self.Memberflag = True
        self.LuckyGroup = 0
        self.LuckyMember = ''
        self.GroupInfo, self.Groups, self.Finished = self.importInfo()
        self.differenceGroup = list(set(self.Groups).difference(set(self.Finished)))
        self.Members = []


    # 初始导入组 和 已选组的信息
    def importInfo(self):
        # 导入所有组
        groups = pd.DataFrame(pd.read_excel(GROUPS))
        finished = pd.DataFrame(pd.read_excel(FINISHED))
        retFinished = list()

        grow, gcol = groups.shape
        if grow == 0:
            tkinter.messagebox.showwarning(title='提示', message='\"' + GROUPS+' \"分组文件中没有数据！！！')
        frow, fcol= finished.shape
        if frow != 0:
            retFinished = finished['组号'].values.tolist()
        group = groups['组号'].values.tolist()
        return groups, group, retFinished


    # 保存抽取的组 和 组员信息
    def saveSelect(self):
        wb = openpyxl.load_workbook(FINISHED)
        # 获取某一特定的工作表
        sheet = wb["Sheet"]

        match = pd.DataFrame(pd.read_excel(MATCH))
        question = match[match['组号'] == self.LuckyGroup].values.tolist()[0][2]
        now = datetime.now()
        strnow = datetime.strftime(now, '%Y-%m-%d %H:%M:%S')

        sheet.append([self.LuckyGroup, question, self.LuckyMember, strnow])
        wb.save(FINISHED)


    # 抽取完成之后进行更新
    def updateDiff(self):
        self.GroupInfo, self.Groups, self.Finished = self.importInfo()
        self.differenceGroup = list(set(self.Groups).difference(set(self.Finished)))


    # 选择 组 的滚动条
    def switch(self):
        self.Groupflag = True
        while self.Groupflag:
            i = random.randint(0, len(self.Groups)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            self.third['text'] = 'Group' + str(self.Groups[i])
            time.sleep(self.speed)


    # 选择 成员 的滚动条
    def switch2(self):
        self.Memberflag = True
        while self.Memberflag:
            i = random.randint(0, len(self.Members)-1)
            self.first2['text'] = self.second2['text']
            self.second2['text'] = self.third2['text']
            self.third2['text'] = str(self.Members[i])
            time.sleep(self.speed)
        # 选组号 开始按钮


    def butStartClick(self):
        t = threading.Thread(target=self.switch)
        t.start()

        # 选组号 结束按钮


    def btnStopClick(self):
        if self.Groupflag == True:
            self.Groupflag = False
            # 随机获取组号
            if len(self.differenceGroup) == 0:
                tkinter.messagebox.showwarning("提示", "所有的组都已经完成！！！")
                return
            index = random.randint(0, len(self.differenceGroup) - 1)
            self.second['text'] = "Group" + str(self.differenceGroup[index])
            self.LuckyGroup = self.differenceGroup[index]
            # 获取组号对应的成员
            members = self.GroupInfo[self.GroupInfo['组号'] == self.LuckyGroup].values.tolist()[0][2:]
            self.Members = members


    # 选成员 开始按钮
    def butStartClick2(self):
        if len(self.Members) != 0:
            t = threading.Thread(target=self.switch2)
            t.start()
        else:
            tkinter.messagebox.showwarning("提示", "请先抽出选择的组！！！")


    # 选成员 结束按钮
    def btnStopClick2(self):
        if self.Memberflag == True:
            self.Memberflag = False
            # self.LuckyMember = self.second2['text']
            index = random.randint(0, len(self.Members)-1)
            self.second2['text'] = self.Members[index]
            self.LuckyMember = self.Members[index]
            # print(self.LuckyMember)
            self.saveSelect()
            # 选一个组号和成员之后需要更新以为选择的组的记录
            self.updateDiff()

    # 刷新所有 flag
    def flush(self):
        self.Groupflag = False
        self.Memberflag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        self.Groupflag = True
        self.Memberflag = True


class Pagetwo_Match(tk.Frame):
    '''第2页'''

    def __init__(self, parent, root):
        super().__init__(parent)
        self.LabelsFrame = tk.LabelFrame(self, text="你们做什么？", padx=25, pady=20)
        self.LabelsFrame.place(x=20, y=35, width=WIDTH/2 -30, height=HEIGHT-100)

        self.first = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame , text='？？？', font=("宋体", 30, "normal"))
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame , text='？？？', font=("宋体", 20, "normal"))
        self.third.pack()


        self.first2 = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.first2.pack()
        self.second2 = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 30, "normal"))
        self.second2['fg'] = 'red'
        self.second2.pack()
        self.third2 = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.third2.pack()

        self.frameBtn = tk.Frame(self.LabelsFrame, width=200, height=50, pady=10, padx=10)
        self.frameBtn.pack()
        # self.frameBtn.place(x=WIDTH/2+20, y=50, width=WIDTH/2 -30, height=HEIGHT-100)

        self.btnStart = tk.Button(self.frameBtn, text='开始', fg='white', bg='green', command=self.butStartClick)
        self.btnStart.place(x=0, y=5, width=45, height=25)
        self.btnEnd = tk.Button(self.frameBtn, text='结束', fg='white', bg='green', command=self.butEndClick)
        self.btnEnd.place(x=50, y=5, width=45, height=25)
        self.btnSave = tk.Button(self.frameBtn, text='SAVE', fg='white', bg='green', command=self.SaveClick)
        self.btnSave.place(x=100, y=5, width=45, height=25)

        self.lframeText = tk.LabelFrame(self, text='匹配信息', pady=10, padx=10)
        self.lframeText.place(x=WIDTH / 2 + 20, y=35, width=WIDTH / 2 - 30, height=HEIGHT - 100)
        self.text = tk.Text(self.lframeText)
        self.text.place(x=0, y=2, width=WIDTH/2-55, height=HEIGHT-150)

        self.speed = 0.02  # 滚动速度
        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions = self.matchInfo()
        # 防止 开始按钮 多次点击
        self.startMatchFlag = True
        # 匹配前控制 滚动条 的滚动信号
        self.flag = True
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = True
        # 匹配时候禁止点击 停止 按钮
        self.showflag = True

        # 组号 --- 题号
        self.Match = {}


    def matchInfo(self):
        # 导入所有组
        groups = pd.DataFrame(pd.read_excel(GROUPS))
        missions = pd.DataFrame(pd.read_excel(MISSIONS))
        mrow, mcol = missions.shape
        if mrow == 0:
            tkinter.messagebox.showwarning("提示", '\"'+MISSIONS+"\" 题目文件中没有信息！")
        g_num = groups['组号'].values.tolist()
        random.shuffle(g_num)
        m_num = missions['题号'].values.tolist()
        random.shuffle(m_num)
        # print(g_num, m_num)
        return groups, missions, g_num, m_num


    # 选择 组/题 的滚动条
    def switch(self):
        # self.flag = True
        while self.flag:
            i = random.randint(0, len(self.Missions)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            self.third['text'] = 'Mission' + str(self.Missions[i])

            j = random.randint(0, len(self.Groups) - 1)
            self.first2['text'] = self.second2['text']
            self.second2['text'] = self.third2['text']
            self.third2['text'] = 'Group' + str(self.Groups[j])
            time.sleep(self.speed)
            if( self.flag == False):
                break


    def show(self):
        if self.showflag == True:
            # 防止多次点击 结束 按钮
            self.showflag = False
            self.text.insert('end', '\n' + "组号\t|\t题号" )
            self.text.insert('end', '\n--------------------------')
            for gro in self.Match.keys():
                self.second['text'] = "组号" + str(gro)
                self.second2['text'] = "题号" + str(self.Match[gro])
                self.text.insert('end', '\n' + str(gro) + "\t|\t" + str(self.Match[gro]))
                time.sleep(0.8)
            self.text.insert('end', '\n\n\t 分组数目：' + str(len(self.Match.keys())))
            # 控制匹配的时候 的滚动
            self.fflag = False
            tkinter.messagebox.showinfo(title="提示", message="匹配完成QAQ")


    def switch2(self):
        # self.fflag = True
        while self.fflag:
            i = random.randint(0, len(self.Missions)-1)
            self.first['text'] = self.third['text']
            # self.second['text'] = self.third['text']
            self.third['text'] = 'Mission' + str(self.Missions[i])

            j = random.randint(0, len(self.Groups) - 1)
            self.first2['text'] = self.third2['text']
            # self.second2['text'] = self.third2['text']
            self.third2['text'] = 'Group' + str(self.Groups[j])
            time.sleep(self.speed)
            if( self.fflag == False):
                break


    # 选组号 开始按钮
    def butStartClick(self):
        if(self.startMatchFlag == True):
            t = threading.Thread(target=self.switch)
            t.start()
            self.startMatchFlag = False

        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions = self.matchInfo()
        for i in range(0, len(self.Groups)):
            self.Match[self.Groups[i]] = self.Missions[i % len(self.Missions)]


    def butEndClick(self):
        self.flag = False
        t = threading.Thread(target=self.show)
        t.start()
        t2 = threading.Thread(target=self.switch2)
        t2.start()


    # 保存匹配结果
    def SaveClick(self):
        if self.flag != False:
            tkinter.messagebox.showinfo(title="提示", message="您还没有进行匹配QAQ")
            return
        if tkinter.messagebox.askyesno(title="提示",message="确定保存分配结果吗？") == True:
            self.save_MatchInfo()


    def save_MatchInfo(self):
        match = pd.DataFrame(columns=('组号', '题号', '题目', '成员1', '成员2', '成员3', '成员4', '备注'))
        for Gnum in range(0, len(self.Groups)):
            name = self.MissionsInfo[self.MissionsInfo['题号'] == self.Missions[Gnum]].values.tolist()[0][1]
            members = self.GroupsInfo[self.GroupsInfo['组号'] == self.Groups[Gnum]].values.tolist()[0][2:]
            newRow = pd.DataFrame({'组号': self.Groups[Gnum], '题号': self.Missions[Gnum], '题目': name,
                                    '成员1': members[0], '成员2': members[1], '成员3': members[2], '成员4': members[3],
                                    '备注': ""}, index=[Gnum+1])
            match = match.append(newRow)
        match.to_excel(MATCH)
        tkinter.messagebox.showinfo(title="提示", message="保存成功")


    def flush(self):
        # 防止 开始按钮 多次点击
        self.startMatchFlag = False
        # 匹配前控制 滚动条 的滚动信号
        self.flag = False
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = False
        # 匹配时候禁止点击 停止 按钮
        self.showflag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        # ===================================
        # 防止 开始按钮 多次点击
        self.startMatchFlag = True
        # 匹配前控制 滚动条 的滚动信号
        self.flag = True
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = True
        # 匹配时候禁止点击 停止 按钮
        self.showflag = True

        self.text.delete(0.0, tkinter.END)


class PageThree_New(tk.Frame):
    '''第3页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        """        
        self.label = tk.Label(self, text="这是第二页", font=LARGE_FONT)
        self.label.pack(pady=10, padx=10)
        self.grlabel = tk.Label(self, text="分组文件名", font=LARGE_FONT)
        self.grlabel.pack()
        self.grentry = tk.Entry(self, show=None, font=('Arial', 14))
        self.grentry.pack()

        self.missionlabel = tk.Label(self, text="题目文件名", font=LARGE_FONT)
        self.missionlabel.pack()
        self.missionentry = tk.Entry(self, show=None, font=('Arial', 14))
        self.missionentry.pack()

        self.matchlabel = tk.Label(self, text="匹配文件名", font=LARGE_FONT)
        self.matchlabel.pack()
        self.matchentry = tk.Entry(self, show=None, font=('Arial', 14))
        self.matchentry.pack()
        """

        self.Confrimbutton = tk.Button(self, text="点击确认开始新学期", fg='white', bg='green', command=self.confrim)
        self.Confrimbutton.pack(padx=50, pady=50)

    def confrim(self):
        """
        group = self.grentry.get()
        mission = self.missionentry.get()
        match = self.matchentry.get()
        if group=='' or mission=='' or match=='':
            tkinter.messagebox.showinfo(title='Hi', message='请填写要创建的文件名称！！！')
            return
        msg = group+mission+match
        """
        if tkinter.messagebox.askyesno(title='提示', message="确定创建文件！开始新学期吗？") == True:
            self.modify()
            tkinter.messagebox.showinfo(title='提示', message='新学期所需的文件groups、missions、finished已经创建在config文件夹中！')
        else:
            tkinter.messagebox.showerror(title='Hi', message='取消操作')


    def getTimeInfo(self):
        nowt = datetime.now()
        timeInfo = datetime.strftime(nowt, '%Y-%m-%d')
        return timeInfo


    def modify(self):
        timeInfo = self.getTimeInfo()
        config = "config/"
        groups ='groups'+timeInfo+".xlsx"
        missions = 'missions'+timeInfo+".xlsx"
        finished = 'finished'+timeInfo+".xlsx"
        with open("config/config.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict['groups'] = groups
            load_dict['mission'] = missions
            load_dict['finished'] = finished
            str = json.dumps(load_dict)
            load_f.seek(0)
            load_f.truncate()
            load_f.write(str)
        # ==================================
        wb1 = openpyxl.Workbook()
        sheet = wb1['Sheet']
        sheet.append(['组号', '组名', '成员1', '成员2', '成员3', '成员4'])
        wb1.save(config+groups)
        # ==================================
        wb2 = openpyxl.Workbook()
        sheet2 = wb2['Sheet']
        sheet2.append(['组号', '题号', '成员', '时间'])
        wb2.save(config+finished)
        # ==================================
        wb3 = openpyxl.Workbook()
        sheet3 = wb3['Sheet']
        sheet3.append(['题号', '题目', '要求', '备注'])
        wb3.save(config+missions)


    def flush(self):
        pass


class PageFour_Luc(tk.Frame):
    '''第4页'''
    def __init__(self, parent, root):
        super().__init__(parent)

        self.LabelsFrame = tk.LabelFrame(self, text="谁是幸运星", padx=25, pady=20)
        self.LabelsFrame.place(x=50, y=50, width=WIDTH-120, height=HEIGHT-100)
        self.first = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 40, "normal"))
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"))
        self.third.pack()
        self.Confrimbutton = tk.Button(self.LabelsFrame, text="Lucky", fg='white', bg='green', command=self.ramdom_peo)
        self.Confrimbutton.pack(padx=50, pady=50)

    def ramdom_peo(self):
        pass


    def flush(self):
        pass


if __name__ == '__main__':
    # 初始化配置
    init()
    # 实例化Application
    app = Application()
    # 主消息循环:
    app.mainloop()