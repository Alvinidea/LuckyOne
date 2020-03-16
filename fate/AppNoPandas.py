from datetime import datetime
import tkinter as tk
import tkinter.font as tf
import time
import tkinter.messagebox
import random
import threading
# import pandas as pd
import json
import os
import openpyxl
from openpyxl import load_workbook
'''
需要初始文件
    groups.xlsx
    finished.xlsx
    missions.xlsx
'''
LARGE_FONT = ("Verdana", 12)
'''
GROUPS = "config/groups.xlsx"
MISSIONS = "config/missions.xlsx"
FINISHED = "config/finished.xlsx"
MATCH = "config/match.xlsx"

配置文件： config.txt
{
"groups":"groups.xlsx",
"match":"match.xlsx",
"mission":"missions.xlsx",
"finished":"finished.xlsx",
"classlist":"classlist.xlsx"
}
'''
# 加载配置信息
def init():
    global config, CLASSNAME
    global GROUPS, MISSIONS, FINISHED, MATCH, WIDTH, CLASSLIST, HEIGHT
    with open("config/config.txt", 'r', encoding='utf-8') as load_f:
        config = "config/"
        load_dict = json.load(load_f)
        CLASSNAME = load_dict['classname']
        GROUPS = config + load_dict['groups']
        MISSIONS = config + load_dict['mission']
        FINISHED = config + load_dict['finished']
        MATCH = config + load_dict['match']
        # 班级名单
        CLASSLIST = config + load_dict['classlist']


    WIDTH = 1000
    HEIGHT = 850
    # 判断 3个文件是否存在
    if os.path.exists(GROUPS) == False:
        tkinter.messagebox.showwarning(title='Warn', message="分组信息不存在！请导入分组信息到config文件夹中！")
    if os.path.exists(MISSIONS) == False:
        tkinter.messagebox.showwarning(title='Warn', message="题目信息不存在！请导入题目信息到config文件夹中！")
    if os.path.exists(CLASSLIST) == False:
        tkinter.messagebox.showwarning(title='Warn', message="花名册信息不存在！请导入班级信息到config文件夹中！")
    if os.path.exists(FINISHED) == False:
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.append(['组号', '题号', '成员', '时间'])
        wb.save(FINISHED)

# excel的操作类
class excel():
    def __init__(self, file, create=True):
        self.file = file
        if create == True:
            self.wb = load_workbook(self.file, data_only=True)
            sheets = self.wb.sheetnames
            self.sheet = sheets[0]
            self.ws = self.wb[self.sheet]
        else:
            self.wb = openpyxl.Workbook()
            sheets = self.wb.sheetnames
            self.sheet = sheets[0]
            self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        for index in range(1, rows+1):
            if self.getCellValue(index, 2) == '' or self.getCellValue(index, 2) is None :
                rows = index-1
                break
        for index in range(1, columns+1):
            if self.getCellValue(1, index) is None:
                columns = index-1
                break
        return rows, columns


    def getRowsNum(self):
        rows = self.ws.max_row
        for index in range(1, rows+1):
            if self.getCellValue(index, 2) == '' or self.getCellValue(index, 2) is None :
                rows = index-1
                break
        return rows


    def getColsNum(self):
        columns = self.ws.max_column
        for index in range(1, columns+1):
            if self.getCellValue(1, index) is None:
                columns = index-1
                break
        return columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    # 获取某行所有值
    def addRow(self, row):
        try:
            self.ws.append(row)
            self.wb.save(self.file)
        except:
            self.ws.append([])
            self.wb.save(self.file)

    # 跟据某列的值获取行数据
    def getRowByColVulue(self, column, value):
        cols = self.getColValues(column)
        for irow in range(1, self.getRowsNum()+1):
            if value == self.getCellValue(irow, column):
                # print(self.getRowValues(irow))
                return self.getRowValues(irow)


# 数据获取类  调用excel中的方法获取数据
class interTool():
    def __init__(self):
        pass

    def groupColInfo(self, exc):
        # 获取存储信息的列号
        # {'成员姓名': [4, 6, 8], '成员学号': [3, 5, 7], '组号': 1, '组名': 2, '组长学号': 3, '组长姓名': 4}
        # 导入所有组
        # exc = excel(GROUPS)
        # print(GROUPS)
        grow, gcol = exc.getRowsClosNum()

        tempName = exc.getRowValues(1)
        # print(tempName)
        count = 0
        group_single = {}
        group_single['成员姓名'] = list()
        group_single['成员学号'] = list()
        for tem in tempName:
            count = count + 1
            if tem is None:
                break
            # 组号/序号/队号  组名/队名    队长姓名/组长姓名  队长学号/组长学号   成员姓名/队员姓名  队员学号/成员学号
            if tem == '组号' or tem == '序号' or tem == '队号':
                group_single['组号'] = count
            if tem == '组名' or tem == '队名':
                group_single['组名'] = count

            if '组长' in tem or '队长' in tem:
                if '姓名' in tem :
                    group_single['组长姓名'] = count
                    group_single['成员姓名'].append(count)
                if '学号' in tem :
                    group_single['组长学号'] = count
                    group_single['成员学号'].append(count)
            if '成员' in tem or '队员' in tem:
                if '姓名' in tem:
                    group_single['成员姓名'].append(count)
                if '学号' in tem:
                    group_single['成员学号'].append(count)

        # print(group_single)
        return group_single


    def missionColInfo(self, exc):
        # 获取存储信息的列号
        row, col = exc.getRowsClosNum()
        tempName = exc.getRowValues(1)
        count = 0
        m_single = {}
        for tem in tempName:
            count = count + 1
            if tem is None:
                break
            # "题号","题目","要求","备注"
            if tem == "题号" or tem == '序号' or tem == '编号':
                m_single["题号"] = count
            if tem == "题目" or tem == "标题":
                m_single["题目"] = count
            if "要求" == tem :
                m_single['要求'] = count
            if '备注' == tem :
                m_single['成员学号'] = count
        return m_single


    #===========================================================================

    # 初始导入组 和 已选组的信息
    def PageZore_init(self):
        # 导入所有组
        exc = excel(MISSIONS)
        # print(GROUPS)
        mrow, mcol = exc.getRowsClosNum()
        # print(grow, gcol)
        nums = exc.getColValues(1)[1:mrow]
        names = exc.getColValues(2)[1:mrow]
        return nums, names, mrow-1


    #===========================================================================


    def PageSign_init(self):
        # 导入所有组
        exc = excel(GROUPS)
        group_single = self.groupColInfo(exc)
        # print(GROUPS)
        grow, gcol = exc.getRowsClosNum()
        # print(grow, gcol)
        # 按照行来存取
        groups = {}
        if grow <= 1:
            tkinter.messagebox.showwarning(title='提示',
                                           message='\"' + GROUPS + ' \"分组文件中没有数据！！！')
        for index in range(1, grow + 1):
            if exc.getCellValue(index, 2) is None:
                break
            tempG = exc.getRowValues(index)
            groups[tempG[0]] = tempG[0:gcol]
        # 获取组号/序号 列表
        # tempName = exc.getRowValues(1)
        tempName = exc.getRowValues(group_single['组号'])
        count = 0
        group = list()
        for tem in tempName:
            count = count + 1
            if tem == '组号' or tem == '序号' or tem == '队号':
                group = exc.getColValues(count)[1:grow]
        # group = exc.getColValues(1)[1:]

        retFinished = self.PageSign_initfinished()

        return groups, group, retFinished, group_single


    def PageSign_initfinished(self):
        fexc = excel(FINISHED)
        frow, fcol = fexc.getRowsClosNum()
        retFinished = list()

        if frow > 1:
            retFinished = fexc.getColValues(1)[1:frow]
        # print(group, retFinished)
        return retFinished

    # 保存抽取的组 和 组员信息
    def PageSign_saveSelect(self, LuckyGroup, LuckyMember):
        wb = openpyxl.load_workbook(FINISHED)
        # 获取某一特定的工作表
        sheet = wb["Sheet"]
        if os.path.exists(MATCH) == False:
            tkinter.messagebox.showwarning(title="提示", message=MATCH+"文件未创建！请进行分组题目1：1匹配")
            return
        #
        exc = excel(MATCH)
        # 2 : 根据 组号（第一列）的值  ，获取该列所在的行
        question = exc.getRowByColVulue(1, LuckyGroup)[1]
        now = datetime.now()
        strnow = datetime.strftime(now, '%Y-%m-%d %H:%M:%S')
        # print([self.LuckyGroup, question, self.LuckyMember, strnow])

        sheet.append([LuckyGroup, question, LuckyMember, strnow])
        wb.save(FINISHED)


    #================================================================================

    def PageMatchInfo(self):
        # 导入所有组
        exc = excel(GROUPS)
        grow, gcol = exc.getRowsClosNum()
        groups = {}
        for index in range(1, grow+1):
            tempG = exc.getRowValues(index)
            groups[tempG[0]] = tempG[0:gcol]
        group = exc.getColValues(1)[1:grow]
        group_single = self.groupColInfo(exc)

        mexc = excel(MISSIONS)
        mission_single = self.missionColInfo(mexc)
        mrow, mcol = mexc.getRowsClosNum()
        missions = {}
        for index in range(1, mrow + 1):
            tempM = mexc.getRowValues(index)
            missions[tempM[0]] = tempM[0:mcol]
        mission = mexc.getColValues(1)[1:mrow]

        if mrow == 0:
            tkinter.messagebox.showwarning("提示", '\"'+MISSIONS+"\" 题目文件中没有信息！")

        random.shuffle(group)
        random.shuffle(mission)
        return groups, missions, group, mission, group_single, mission_single

    def PageMatch_saveInfo(self, GroupsInfo,  Groups, group_single, MissionsInfo, mission_single, Match):
        exc = excel(MATCH, False)
        row = ['组号','组名', '题号', '题目']\
              +['成员'+str(index) for index in range(1, len(group_single['成员姓名'])+1)]\
              +['备注']
        # print(row)
        # row = ['组号','组名', '题号', '题目', '成员1', '成员2', '成员3', '成员4', '备注']
        exc.addRow(row)
        for Gnum in range(0, len(Groups)):
            title_col = mission_single['题目']
            MissionNum = Match[Groups[Gnum]]
            mname = MissionsInfo[MissionNum][title_col-1]

            # members = self.GroupsInfo[self.Groups[Gnum]][2:]
            names_col = group_single['成员姓名']
            row = GroupsInfo[Groups[Gnum]]
            gname = row[group_single['组名']-1]
            members = [row[index-1] for index in names_col]
            newRow = [Groups[Gnum], gname, MissionNum, mname]+members+[""]
            """newRow = [self.Groups[Gnum], self.Missions[Gnum], name,
                      members[0], members[1], members[2], members[3],
                      ""]"""
            exc.addRow(newRow)
        tkinter.messagebox.showinfo(title="提示", message="保存成功")


    #================================================================================

    def PageNew_init(self):
        with open("config/config.txt", 'r', encoding='utf-8') as load_f:
            config = "config/"
            listc = []
            load_dict = json.load(load_f)
            # print(load_dict)
            listc.append(config + load_dict['groups'])
            listc.append(config + load_dict['mission'])
            listc.append(config + load_dict['finished'])
            listc.append(config +load_dict['match'])
            # 班级名单
            """ 
            CLASSLIST = config + load_dict['classlist']
            ClASSNAME = load_dict['classname']
            """
            return listc

    # 创建新班级的配置文件
    # config/ className / config file


    def PageNew_modify(self, entryClassName):
        config = "config/"
        # className = "Test"
        className = entryClassName
        if os.path.exists(config+className) == True:
            tkinter.messagebox.showwarning(title="提示", message=className+" 文件夹已经存在，请重新输入班级名称！")
            return
        # 创建文件夹
        os.makedirs(config+className)
        # timeInfo = self.getTimeInfo()
        groups = className+'/groups'+".xlsx"
        missions = className+'/missions'+".xlsx"
        finished = className+'/finished'+".xlsx"
        classlist = className+'/classlist'+".xlsx"

        # 更改当前配置
        """with open("config/config.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict['groups'] = groups
            load_dict['mission'] = missions
            load_dict['finished'] = finished
            load_dict['match'] = className + '/match.xlsx'
            load_dict['classlist'] = classlist
            load_dict['classname'] = className

            # configinfo['match'] = className + '/match.xlsx'=============================
            str = json.dumps(load_dict)
            load_f.seek(0)
            load_f.truncate()
            load_f.write(str)"""

        # ==================================
        wb1 = openpyxl.Workbook()
        sheet = wb1['Sheet']
        sheet.append(['组号', '组名', '组长学号', '组长姓名', '成员1学号', '成员1姓名','成员2学号', '成员2姓名'])
        wb1.save(config+groups)
        # ==================================
        wb2 = openpyxl.Workbook()
        sheet2 = wb2['Sheet']
        sheet2.append(['组号', '题号', '成员', '时间'])
        wb2.save(config+finished)
        # ==================================
        wb3 = openpyxl.Workbook()
        sheet3 = wb3['Sheet']
        sheet3.append(['题号', '题目', '要求', '备注'])
        wb3.save(config+missions)
        # ==================================
        wb3 = openpyxl.Workbook()
        sheet3 = wb3['Sheet']
        sheet3.append(['姓名', 'name', 'Name'])
        wb3.save(config + classlist)

        configinfo = {}
        configinfo['groups'] = groups
        configinfo['mission'] = missions
        configinfo['finished'] = finished
        configinfo['classlist'] = classlist
        configinfo['match'] = className+'/match.xlsx'
        configinfo['classname'] = className
        self.PageNew_addConfigToClassConfig(className, configinfo)


    def PageNew_addConfigToClassConfig(self, configName,configInfo):
        with open("config/ClassConfig.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict[configName] = configInfo
            str = json.dumps(load_dict)
            load_f.seek(0)
            # 清空之前的信息
            load_f.truncate()
            load_f.write(str)

    # ================================================================================

    def PageLuc_initClassInfo(self):
        if os.path.exists(CLASSLIST) == False:
            # tkinter.messagebox.showwarning(title='提示！',message=CLASSLIST+"文件不存在，无法实现单人点名操作" )
            return dict()
        exc = excel(CLASSLIST)
        rown = exc.getColsNum()
        classInfo = {}
        for col in range(1, rown+1):
            tempCol = exc.getColValues(col)
            classInfo[tempCol[0]] = tempCol[1:]
        return classInfo


class Application(tk.Tk):
    '''
    多页面测试程序
        界面与逻辑分离
    '''
    def __init__(self):
        super().__init__()

        self.title("Fate")
        self.geometry('1000x850')

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        menubar = tk.Menu(self)
        filemenu = tk.Menu(menubar, tearoff=0)
        newmenu = tk.Menu(menubar, tearoff=0)

        menubar.add_cascade(label='决定命运', menu=filemenu)
        menubar.add_cascade(label='新学期', menu=newmenu)

        filemenu.add_command(label="显示题目", command=self.showmission)
        filemenu.add_command(label="抽签选组", command=self.chouqian)
        filemenu.add_command(label="分配题目", command=self.xuanti)
        filemenu.add_command(label="随机抽人", command=self.luckypeo)

        newmenu.add_command(label="选择班级", command=self.selectClass)
        newmenu.add_command(label="新学期", command=self.xinxueqi)
        newmenu.add_command(label="刷新", command=self.flush)
        self.config(menu=menubar)
        self.frames = {}
        for F in (ShowMission_Zero, StartPage_Sign, Pagetwo_Match, PageThree_New, PageFour_Luc, PageFive_selectClass):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
            # 四个页面的位置都是 grid(row=0, column=0), 位置重叠，只有最上面的可见！！
        self.show_frame(ShowMission_Zero)


    # 展示框架
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()  # 切换，提升当前 tk.Frame z轴顺序（使可见）！！此语句是本程序的点睛之处


    def showmission(self):
        self.show_frame(ShowMission_Zero)

    def chouqian(self,):
        self.show_frame(StartPage_Sign)


    def xuanti(self):
        self.show_frame(Pagetwo_Match)


    def xinxueqi(self):
        self.show_frame(PageThree_New)

    def selectClass(self):
        self.show_frame(PageFive_selectClass)


    def luckypeo(self):
        self.show_frame(PageFour_Luc)


    def flush(self):
        init()
        for F in (ShowMission_Zero, StartPage_Sign, Pagetwo_Match, PageThree_New, PageFour_Luc, PageFive_selectClass):
            frame = self.frames[F]
            frame.flush()
        tkinter.messagebox.showinfo(title='提示', message="刷新成功！")


class ShowMission_Zero(tk.Frame):
    '''主1页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.tool = interTool()
        self.LabelsFrame = tk.LabelFrame(self, text="题目信息", padx=35, pady=35)
        # self.LabelsFrame1.pack()
        self.LabelsFrame.place(x=10, y=10, width=WIDTH-20, height=HEIGHT-20)
        self.nums, self.names, self.mrow = self.tool.PageZore_init()

        self.text = tk.Text(self.LabelsFrame)
        self.text.place(x=0, y=2, width=WIDTH - 200, height=HEIGHT - 150)
        self.textshow()
        """        
        self.num_names = []
        strs = ''
        for index in range(0, self.mrow):
            stri = str(self.nums[index]) + '\t\t' + str(self.names[index])+ '|\n'
            strs = strs +"|\t"+ stri
        self.num_name = tk.Label(self.LabelsFrame,text=strs, height=HEIGHT-30 ,font=("宋体", 16, "normal"))
        self.num_name.pack()"""


    def textshow(self):
        self.text.delete(0.0, tkinter.END)
        if len(self.nums) == 0:
            return
        self.text.tag_config('tag', font=tf.Font(family='微软雅黑', size=14))
        self.showflag = False
        self.text.insert(tk.END, "\t题号\t\t\t|\t题目", 'tag')
        self.text.insert(tk.END, '\n\t--------------------------------------', 'tag')
        for index in range(0, self.mrow):
            info = '\n\t' + str(self.nums[index]) + '\t\t\t|\t' + str(self.names[index])
            self.text.insert(tk.END, info, 'tag')

    def flush(self):
        self.nums, self.names, mrow = self.tool.PageZore_init()
        index = 0
        self.textshow()


class StartPage_Sign(tk.Frame):
    '''主1页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.tool = interTool()
        self.label = tk.Label(self, text="谁是幸运儿！！！", font=LARGE_FONT,bg='green', fg='white')
        # self.label.pack(pady=10, padx=10)
        self.label.place(x=WIDTH/2-100, y=10, width=200, height=25)
        # -------------------------------------------- 选组
        self.LabelsFrame1 = tk.LabelFrame(self, text="哪组会中奖？", padx=35, pady=35)
        # self.LabelsFrame1.pack()
        self.LabelsFrame1.place(x=20, y=50, width=WIDTH/2 -30, height=HEIGHT-200)
        self.first = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 20, "normal"), pady=15)
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 30, "normal"), pady=25)
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame1, text='？？？', font=("宋体", 20, "normal"), pady=15)
        self.third.pack()

        self.frame1 = tk.Frame(self.LabelsFrame1, width=120, height=50, pady=10, padx=10)
        self.frame1.pack()
        self.btnStart = tk.Button(self.frame1, text='开始', bg='green', fg='white', command=self.butStartClick)
        self.btnStart.place(x=0, y=5, width=45, height=25)
        self.butStop = tk.Button(self.frame1, text='停止',  bg='green', fg='white', command=self.btnStopClick)
        self.butStop.place(x=50, y=5, width=45, height=25)
        # -------------------------------------------- 选人
        self.LabelsFrame2 = tk.LabelFrame(self, text="哪位同学展示？", padx=35, pady=35)
        # self.LabelsFrame2.pack()
        self.LabelsFrame2.place(x=WIDTH/2+20, y=50, width=WIDTH/2 -30, height=HEIGHT-200)
        self.first2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 20, "normal"), pady=15)
        self.first2.pack()

        self.second2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 30, "normal"), pady=25)
        self.second2['fg'] = 'red'
        self.second2.pack()

        self.third2 = tk.Label(self.LabelsFrame2, text='???', font=("宋体", 20, "normal"), pady=15)
        self.third2.pack()
        self.frame2 = tk.Frame(self.LabelsFrame2, width=120, height=50, pady=10, padx=10)
        self.frame2.pack()
        self.btnStart2 = tk.Button(self.frame2, text='开始', bg='green', fg='white', command=self.butStartClick2)
        self.btnStart2.place(x=0, y=5, width=45, height=25)
        self.butStop2 = tk.Button(self.frame2,  text='停止',  bg='green', fg='white', command=self.btnStopClick2)
        self.butStop2.place(x=50, y=5, width=45, height=25)
        # LabelFrame
        self.speed = 0.02           # 滚动速度
        self.Groupflag = True
        self.Memberflag = True

        self.LuckyGroup = 0
        self.LuckyMember = ''
        self.GroupInfo, self.Groups, self.Finished, self.group_single = self.tool.PageSign_init()
        self.differenceGroup = list(set(self.Groups).difference(set(self.Finished)))
        self.Members = []


    # 抽取完成之后进行更新
    def updateDiff(self):
        self.GroupInfo, self.Groups, self.Finished, self.group_single = self.tool.PageSign_init()
        self.differenceGroup = list(set(self.Groups).difference(set(self.Finished)))


    # 选择 组 的滚动条
    def switch(self):
        self.Groupflag = True
        while self.Groupflag:
            i = random.randint(0, len(self.Groups)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            showname = str(self.Groups[i]) +": "+str(self.GroupInfo[self.Groups[i]][self.group_single['组名']-1])
            self.third['text'] =  showname
            time.sleep(self.speed)


    # 选择 成员 的滚动条
    def switch2(self):
        self.Memberflag = True
        while self.Memberflag:
            i = random.randint(0, len(self.Members)-1)
            self.first2['text'] = self.second2['text']
            self.second2['text'] = self.third2['text']
            self.third2['text'] = str(self.Members[i])
            time.sleep(self.speed)
        # 选组号 开始按钮


    def butStartClick(self):
        if len(self.Groups) == 0:
            tkinter.messagebox.showwarning(title='Warn', message="分组文件只有列名没有数据！！！")
            return
        t = threading.Thread(target=self.switch)
        t.start()

        # 选组号 结束按钮


    def btnStopClick(self):
        if self.Groupflag == True:
            self.Groupflag = False
            # 随机获取组号
            if len(self.differenceGroup) == 0:
                tkinter.messagebox.showwarning("提示", "所有的组都已经完成！！！")
                return
            index = random.randint(0, len(self.differenceGroup) - 1)
            self.LuckyGroup = self.differenceGroup[index]
            showname = str(self.LuckyGroup) +": "+ self.GroupInfo[self.LuckyGroup][self.group_single['组名'] - 1]
            self.second['text'] = showname

            # 获取组号对应的成员
            # members = self.GroupInfo[self.GroupInfo['组号'] == self.LuckyGroup].values.tolist()[0][2:]
            # print(self.GroupInfo, self.LuckyGroup)
            ###############################################################
            names = self.group_single['成员姓名']
            row = self.GroupInfo[self.LuckyGroup]
            members = [row[index-1] for index in names]
            self.Members = members


    # 选成员 开始按钮
    def butStartClick2(self):
        if len(self.Members) != 0:
            t = threading.Thread(target=self.switch2)
            t.start()
        else:
            tkinter.messagebox.showwarning("提示", "请先抽出选择的组！！！")


    # 选成员 结束按钮
    def btnStopClick2(self):
        if self.Memberflag == True:
            self.Memberflag = False
            # self.LuckyMember = self.second2['text']
            index = random.randint(0, len(self.Members)-1)
            # print(self.Members)
            random.shuffle(self.Members)
            self.second2['text'] = self.Members[index]
            self.LuckyMember = self.Members[index]
            # print(self.LuckyMember)
            self.tool.PageSign_saveSelect(self.LuckyGroup, self.LuckyMember)
            # 选一个组号和成员之后需要更新以为选择的组的记录
            self.updateDiff()

    # 刷新所有 flag
    def flush(self):
        self.Groupflag = False
        self.Memberflag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        self.Groupflag = True
        self.Memberflag = True

        self.LuckyGroup = 0
        self.LuckyMember = ''
        self.updateDiff()
        self.Members = []


class Pagetwo_Match(tk.Frame):
    '''第2页'''

    def __init__(self, parent, root):
        super().__init__(parent)
        self.tool = interTool()
        self.LabelsFrame = tk.LabelFrame(self, text="你们做什么？", padx=25, pady=20)
        self.LabelsFrame.place(x=20, y=35, width=WIDTH/2 -30, height=HEIGHT-100)
        self.LabelsFrame_1 = tk.LabelFrame(self.LabelsFrame, text="题目", padx=25, pady=20)
        self.LabelsFrame_1.place(x=5, y=5, width=WIDTH/2 -80, height=HEIGHT/2-200)
        self.first = tk.Label(self.LabelsFrame_1, text='？？？', font=("宋体", 20, "normal"), pady=15)
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame_1 , text='？？？', font=("宋体", 30, "normal"), pady=15)
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame_1 , text='？？？', font=("宋体", 20, "normal"), pady=15)
        self.third.pack()

        self.LabelsFrame_2 = tk.LabelFrame(self.LabelsFrame, text="队伍", padx=25, pady=20)
        self.LabelsFrame_2.place(x=5, y=HEIGHT/2-180, width=WIDTH/2 -80, height=HEIGHT/2-200)
        self.first2 = tk.Label(self.LabelsFrame_2, text='？？？', font=("宋体", 20, "normal"), pady=15)
        self.first2.pack()
        self.second2 = tk.Label(self.LabelsFrame_2, text='？？？', font=("宋体", 30, "normal"), pady=15)
        self.second2['fg'] = 'red'
        self.second2.pack()
        self.third2 = tk.Label(self.LabelsFrame_2, text='？？？', font=("宋体", 20, "normal"), pady=15)
        self.third2.pack()

        self.frameBtn = tk.Frame(self.LabelsFrame, width=200, height=50, pady=10, padx=10)
        self.frameBtn.place(x=WIDTH/4 -150, y=HEIGHT-320, width=WIDTH/2 -250, height=80)
        # self.frameBtn.place(x=WIDTH/2+20, y=50, width=WIDTH/2 -30, height=HEIGHT-100)

        self.btnStart = tk.Button(self.frameBtn, text='开始', fg='white', bg='green', command=self.butStartClick)
        self.btnStart.place(x=0, y=5, width=45, height=25)
        self.btnEnd = tk.Button(self.frameBtn, text='结束', fg='white', bg='green', command=self.butEndClick)
        self.btnEnd.place(x=50, y=5, width=45, height=25)
        self.btnSave = tk.Button(self.frameBtn, text='SAVE', fg='white', bg='green', command=self.SaveClick)
        self.btnSave.place(x=100, y=5, width=45, height=25)

        self.lframeText = tk.LabelFrame(self, text='匹配信息', pady=10, padx=10)
        self.lframeText.place(x=WIDTH / 2 + 20, y=35, width=WIDTH / 2 - 30, height=HEIGHT - 100)
        self.text = tk.Text(self.lframeText)
        self.text.place(x=0, y=2, width=WIDTH/2-55, height=HEIGHT-150)

        self.speed = 0.02  # 滚动速度
        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions, self.group_single, self.mission_single \
            = self.tool.PageMatchInfo()
        # 组号 --- 题号
        self.Match = {}
        # 防止 开始按钮 多次点击
        self.startMatchFlag = True
        # 匹配前控制 滚动条 的滚动信号
        self.flag = True
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = True
        # 匹配时候禁止点击 停止 按钮
        self.showflag = True


    # 选择 组/题 的滚动条
    def switch(self):
        # self.flag = True
        while self.flag:
            i = random.randint(0, len(self.Missions)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            showinfo_m = str(self.Missions[i]) + ": " + str(self.MissionsInfo[self.Missions[i]][self.mission_single['题目'] - 1])
            self.third['text'] = showinfo_m

            j = random.randint(0, len(self.Groups) - 1)
            self.first2['text'] = self.second2['text']
            self.second2['text'] = self.third2['text']
            # print(type(self.Groups[j]))
            showinfo = str(self.Groups[j]) +": "+ str(self.GroupsInfo[self.Groups[j]][self.group_single['组名']-1])
            self.third2['text'] = showinfo
            time.sleep(self.speed)
            if( self.flag == False):
                break


    def show(self):
        if self.showflag == True:
            # 防止多次点击 结束 按钮
            self.text.tag_config('tag', font=tf.Font(family='微软雅黑', size=14))
            self.showflag = False
            self.text.insert(tk.END,  "分组信息\t\t|\t题号" ,'tag')
            self.text.insert(tk.END, '\n--------------------------','tag')
            for gro in self.Match.keys():
                showinfo_m = str(self.Match[gro]) + ": " + str(self.MissionsInfo[self.Match[gro]][self.mission_single['题目'] - 1])
                showinfo_g = str(gro) + ": " + str(self.GroupsInfo[gro][self.group_single['组名'] - 1])
                self.second2['text'] = showinfo_g
                self.second['text'] = showinfo_m
                self.text.insert(tk.END, '\n' + showinfo_g + "\t\t|\t" + showinfo_m, 'tag')
                time.sleep(0.8)
            self.text.insert(tk.END, '\n\n\t 分组数目：' + str(len(self.Match.keys())), 'tag')
            # 控制匹配的时候 的滚动
            self.fflag = False
            tkinter.messagebox.showinfo(title="提示", message="匹配完成QAQ")


    def switch2(self):
        # self.fflag = True
        while self.fflag:
            i = random.randint(0, len(self.Missions)-1)
            self.first['text'] = self.third['text']
            # self.second['text'] = self.third['text']
            showinfo_m = str(self.Missions[i]) + ": " + \
                         str(self.MissionsInfo[self.Missions[i]][self.mission_single['题目'] - 1])
            self.third['text'] = showinfo_m

            j = random.randint(0, len(self.Groups) - 1)
            self.first2['text'] = self.third2['text']
            # self.second2['text'] = self.third2['text']
            showinfo = str(self.Groups[j]) + ": " + \
                       str(self.GroupsInfo[self.Groups[j]][self.group_single['组名'] - 1])
            self.third2['text'] = showinfo
            time.sleep(self.speed)
            if( self.fflag == False):
                break


    # 选组号 开始按钮
    def butStartClick(self):
        if len(self.Groups) == 0 or len(self.Missions) == 0:
            tkinter.messagebox.showwarning(title='Warn', message="分组文件或题目文件只有列名没有数据！！！")
            return
        if(self.startMatchFlag == True):
            t = threading.Thread(target=self.switch)
            t.start()
            self.startMatchFlag = False

        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions , self.group_single, self.mission_single\
            = self.tool.PageMatchInfo()
        for i in range(0, len(self.Groups)):
            self.Match[self.Groups[i]] = self.Missions[i % len(self.Missions)]


    def butEndClick(self):
        self.flag = False
        t = threading.Thread(target=self.show)
        t.start()
        t2 = threading.Thread(target=self.switch2)
        t2.start()


    # 保存匹配结果
    def SaveClick(self):
        if self.flag != False:
            tkinter.messagebox.showinfo(title="提示", message="您还没有进行匹配QAQ")
            return
        if tkinter.messagebox.askyesno(title="提示",message="确定保存分配结果吗？") == True:
            self.tool.PageMatch_saveInfo(self.GroupsInfo,  self.Groups, self.group_single,
                                         self.MissionsInfo, self.mission_single,
                                         self.Match)


    def flush(self):
        # 防止 开始按钮 多次点击
        self.startMatchFlag = False
        # 匹配前控制 滚动条 的滚动信号
        self.flag = False
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = False
        # 匹配时候禁止点击 停止 按钮
        self.showflag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        # ===================================
        # 防止 开始按钮 多次点击
        self.startMatchFlag = True
        # 匹配前控制 滚动条 的滚动信号
        self.flag = True
        # 匹配时候控制 滚动条 的滚动信号
        self.fflag = True
        # 匹配时候禁止点击 停止 按钮
        self.showflag = True


        self.text.delete(0.0, tkinter.END)

        self.GroupsInfo, self.MissionsInfo, self.Groups, self.Missions, self.group_single, self.mission_single = self.tool.PageMatchInfo()
        # 组号 --- 题号
        self.Match = {}
        # print("2")


class PageThree_New(tk.Frame):
    '''第3页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.tool = interTool()
        list_label = self.tool.PageNew_init()
        self.LabelFrame = tk.LabelFrame(self, text="配置文件信息", pady=50, width=WIDTH-100, height=HEIGHT-180)
        self.LabelFrame.pack()
        self.grlabel = tk.Label(self.LabelFrame, text=list_label[0], font=("宋体", 16, "normal"))
        self.grlabel.pack()
        self.missionlabel = tk.Label(self.LabelFrame, text=list_label[1], font=("宋体", 16, "normal"))
        self.missionlabel.pack()
        self.matchlabel = tk.Label(self.LabelFrame, text=list_label[2], font=("宋体", 16, "normal"))
        self.matchlabel.pack()
        self.finishedlabel = tk.Label(self.LabelFrame, text=list_label[3], font=("宋体", 16, "normal"))
        self.finishedlabel.pack()

        self.frame = tk.Frame(self, width= WIDTH-20, height=50)
        self.frame.pack(padx=50, pady=10)
        self.entryLabel = tk.Label(self.frame, text="班级名称：",font=("宋体", 12, "normal"))
        self.entryLabel.place(x= (WIDTH-20)/2-200, y=5)
        self.entryClassName = tk.Entry(self.frame)
        self.entryClassName.place(x=(WIDTH-20)/2-60, y=5)
        self.Confrimbutton = tk.Button(self, text="点击确认开始新学期", fg='white', bg='green', command=self.confrim)
        self.Confrimbutton.pack(padx=50, pady=25)



    def confrim(self):
        if tkinter.messagebox.askyesno(title='提示', message="确定创建文件！开始新学期吗？") == True:
            className = self.entryClassName.get()
            if className == "":
                tkinter.messagebox.showwarning(title="提示", message="请输入班级名称！")
            else:
                # self.modify()
                self.tool.PageNew_modify(self.entryClassName.get())
                tkinter.messagebox.showinfo(title='提示', message='新学期所需的文件groups、missions、finished已经创建在config文件夹中！')
        else:
            tkinter.messagebox.showerror(title='Hi', message='取消操作')


    def getTimeInfo(self):
        nowt = datetime.now()
        timeInfo = datetime.strftime(nowt, '%Y-%m-%d')
        return timeInfo


    def flush(self):
        list_label = self.tool.PageNew_init()
        self.grlabel['text'] = list_label[0]
        self.missionlabel['text'] = list_label[1]
        self.matchlabel['text'] = list_label[2]
        self.finishedlabel['text'] = list_label[3]
        #B  print("3")


class PageFour_Luc(tk.Frame):
    '''第4页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.tool = interTool()
        self.LabelsFrame = tk.LabelFrame(self, text="谁是幸运星", padx=25, pady=20)
        self.LabelsFrame.place(x=50, y=50, width=WIDTH-120, height=HEIGHT-100)
        self.first = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"), pady=30)
        self.first.pack()
        self.second = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 50, "normal"), pady=30)
        self.second['fg'] = 'red'
        self.second.pack()
        self.third = tk.Label(self.LabelsFrame, text='？？？', font=("宋体", 20, "normal"), pady=30)
        self.third.pack()
        self.Confrimbutton = tk.Button(self.LabelsFrame, text="幸运星", fg='white', bg='green', command=self.ramdom_peo)
        self.Confrimbutton.pack(padx=25, pady=10)
        self.Repeatbutton = tk.Button(self.LabelsFrame, text="重选", fg='white', bg='green', command=self.repeat)
        self.Repeatbutton.pack(padx=25, pady=10)

        self.speed = 0.02
        self.flag = True
        self.ClassInfo = self.tool.PageLuc_initClassInfo()
        self.NameList = self.getNameList()
        self.NumList = self.getNumList()


    def getNameList(self):
        for colName in self.ClassInfo.keys():
            if colName == "Name" or colName == "姓名" or colName=="name":
                return self.ClassInfo[colName]
        else:
            return list()

    def getNumList(self):
        for colName in self.ClassInfo.keys():
            if colName == "学号" or colName == "编号" or colName == "工号":
                return self.ClassInfo[colName]
        else:
            return list()

    # 选择 组 的滚动条
    def switch(self):
        while self.flag:
            i = random.randint(0, len(self.NameList)-1)
            self.first['text'] = self.second['text']
            self.second['text'] = self.third['text']
            info = str(self.NumList[i]) + str(self.NameList[i])
            self.third['text'] =  info
            time.sleep(self.speed)
            if self == False:
                break

    def threadingStop(self):
        time.sleep(3)
        self.flag = False

    def ramdom_peo(self):
        if len(self.NameList) == 0:
            tkinter.messagebox.showwarning(title='Warn', message="花名册不存在")
            return
        t = threading.Thread(target=self.switch)
        t.start()
        t1 = threading.Thread(target=self.threadingStop)
        t1.start()


    def repeat(self):
        self.flag = True
        self.ramdom_peo()


    def flush(self):
        self.flag = False
        # 花费时间，防止太快转换没有关闭其他线程
        for i in range(1, 100):
            pass
        self.flag = True
        # print("4")
        self.ClassInfo = self.tool.PageLuc_initClassInfo()
        self.NameList = self.getNameList()
        self.NumList = self.getNumList()


"""
班级信息配置分为默认配置 和 选择配置

默认配置的信息：config.txt
选择配置的信息：ClassConfig.txt
"""


class PageFive_selectClass(tk.Frame):
    '''第4页'''
    def __init__(self, parent, root):
        super().__init__(parent)
        self.LabelsFrame = tk.LabelFrame(self, text="班级选择", padx=25, pady=20)
        self.LabelsFrame.place(x=50, y=50, width=WIDTH-120, height=HEIGHT-100)
        self.classes, self.configs = self.initClassInfo()
        self.Lebels = list()
        self.Buttons = list()
        self.Buttons2 = list()
        self.frames = list()
        for clas in self.classes:
            frame = tk.Frame(self.LabelsFrame, width=WIDTH-50, height=50)
            # label = tk.Label(frame, text=clas, font=LARGE_FONT)
            button = tk.Button(frame, text=clas, font=LARGE_FONT, fg='white', bg='green',
                               command=lambda arg=clas: self.selectClass(arg))
            # button2 = tk.Button(frame, text="Del "+clas, fg='white',bg='red', font=LARGE_FONT,
            #                   command=lambda arg=clas: self.deleteClass(arg))
            frame.pack(padx=20, pady=2)
            # label.place(x=5, y=10)
            button.place(x=50, y=10)
            # button2.place(x=150, y=10)
            self.frames.append(frame)
            # self.Lebels.append(label)
            self.Buttons.append(button)
            # self.Buttons2.append(button2)


    def initClassInfo(self):
        with open("config/ClassConfig.txt", 'r', encoding='utf-8') as load_f:
            config = "config/"
            # load_dict = json.load(load_f, strict = False)
            load_dict = json.load(load_f)
            # classes = load_dict.keys()
            # configs = load_dict
            return load_dict.keys(), load_dict


    def getConfigByClass(self, cls):
        return self.configs[cls]


    def reSetMemory(self, config):
        configPath = "config/"
        GROUPS = configPath+config['groups']
        MISSIONS = configPath+config['mission']
        FINISHED = configPath+config['finished']
        MATCH = configPath+config['match']
        CLASSNAME= config['classname']
        # 班级名单
        CLASSLIST = configPath+config['classlist']


    def reSetConfig(self, config):
        # print(config)
        # self.reSetMemory(config)
        with open("config/config.txt", 'r+', encoding='utf-8') as load_f:
            load_dict = json.load(load_f)
            load_dict['groups'] = config['groups']
            load_dict['mission'] = config['mission']
            load_dict['finished'] = config['finished']
            load_dict['match'] = config['match']
            load_dict['classlist'] = config['classlist']
            load_dict['classname'] = config['classname']
            str = json.dumps(load_dict)
            load_f.seek(0)
            load_f.truncate()
            load_f.write(str)


    def selectClass(self, cls):
        ret = tkinter.messagebox.askyesno(title='提示',
                                    message="确定重新设置班级吗？" + cls+"信息吗？")
        if ret != True:
            return
        # 根据班名获取班级配置信息
        config = self.getConfigByClass(cls)
        # 重新设置默认配置信息
        self.reSetConfig(config)
        # 提示 之后需要刷新（将内存信息刷新）
        tkinter.messagebox.showinfo(title='操作成功！',
                                    message="班级配置信息已经重新设置！\n当前班级为："+cls)

    def deleteClass(self, cls):
        ret = tkinter.messagebox.askyesno(title='提示',
                                    message="确定删除" + cls+"信息吗？")
        if ret != True:
            return
        # 根据班名获取班级配置信息
        config = self.getConfigByClass(cls)
        # 重新设置默认配置信息
        self.reSetConfig(config)
        # 提示 之后需要刷新（将内存信息刷新）
        tkinter.messagebox.showinfo(title='操作成功！',
                                    message="班级配置信息已经重新设置！\n当前班级为："+cls)

    def flush(self):
        pass

        # print("5")


if __name__ == '__main__':
    # 初始化配置
    init()
    # 实例化Application
    app = Application()
    # 主消息循环:
    app.mainloop()